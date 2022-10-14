import glob, os

from openpyxl import load_workbook

from contact_parser import get_contacts

files_dir = "files"

os.chdir(files_dir)
for file in glob.glob("*.xlsx"):
    filepath = file
    wb = load_workbook(filepath)
    sheet = wb.active
    rows = sheet.max_row

    for i in range(2, rows + 1):
        try:
            if sheet[f'F{i}'].value:
                email, tel, site = get_contacts(sheet[f'F{i}'].value)
                if email:
                    sheet.cell(row=i, column=14).value = email
                if tel:
                    sheet.cell(row=i, column=13).value = tel
                if site:
                    sheet.cell(row=i, column=15).value = site
        except:
            pass

    wb.save(f'result-{file}')
